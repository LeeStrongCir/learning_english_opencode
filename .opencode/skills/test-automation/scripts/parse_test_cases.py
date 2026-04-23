#!/usr/bin/env python3
"""Parse Excel/CSV test case files into structured JSON for code generation."""

import argparse
import csv
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Column name mappings (standard -> possible variants)
COLUMN_MAPPINGS = {
    "case_id": ["用例id", "case id", "tc#", "编号", "case_id", "id", "用例编号"],
    "title": ["用例标题", "title", "标题", "summary", "用例名称", "name"],
    "precondition": ["前置条件", "precondition", "前置", "前提条件", "preconditions", "前提"],
    "steps": ["测试步骤", "steps", "操作步骤", "步骤描述", "step", "操作说明"],
    "expected": ["预期结果", "expected", "期望结果", "expected result", "预期", "期望"],
    "priority": ["优先级", "priority", "p0/p1/p2", "级别"],
    "test_type": ["测试类型", "type", "类型", "test type", "分类"],
    "test_data": ["测试数据", "test data", "输入数据", "input data", "数据"],
    "remarks": ["备注", "remarks", "note", "注释", "说明"],
}

# Keywords for auto-classification
API_KEYWORDS = [
    "接口", "api", "请求", "post", "get", "put", "delete", "patch",
    "响应", "状态码", "token", "鉴权", "认证", "sdk", "调用",
    "http", "rest", "endpoint", "参数", "header", "body", "json",
    "返回", "返回值", "接口文档", "swagger",
]

WEBUI_KEYWORDS = [
    "点击", "输入", "页面", "按钮", "表单", "弹窗", "下拉",
    "跳转", "可见", "文本", "链接", "选择", "勾选", "滚动",
    "拖拽", "悬停", "右键", "双击", "打开", "关闭", "刷新",
    "导航", "菜单", "标签页", "对话框", "提示", "列表", "表格",
    "搜索框", "输入框", "上传", "下载", "截图", "元素",
]


def normalize_column_name(header):
    """Normalize column header to find matching standard field."""
    header_lower = header.strip().lower()
    for standard, variants in COLUMN_MAPPINGS.items():
        if header_lower in [v.lower() for v in variants]:
            return standard
    return None


def classify_case(case):
    """Auto-classify test case as api, webui, or mixed based on content."""
    # If test_type column exists and is explicit, use it
    test_type = case.get("test_type", "").strip().lower()
    if test_type:
        if "api" in test_type or "接口" in test_type:
            return "api"
        if "ui" in test_type or "web" in test_type or "页面" in test_type:
            return "webui"
        if "mixed" in test_type or "混合" in test_type or "e2e" in test_type:
            return "mixed"

    # Auto-classify based on content
    content = " ".join([
        case.get("title", ""),
        case.get("precondition", ""),
        case.get("steps", ""),
        case.get("expected", ""),
    ]).lower()

    has_api = any(kw in content for kw in API_KEYWORDS)
    has_webui = any(kw in content for kw in WEBUI_KEYWORDS)

    if has_api and has_webui:
        return "mixed"
    elif has_api:
        return "api"
    elif has_webui:
        return "webui"
    else:
        # Default: try to infer from expected result patterns
        if any(kw in content for kw in ["status_code", "status code", "200", "404", "500"]):
            return "api"
        return "webui"


def parse_csv(filepath):
    """Parse CSV test case file."""
    cases = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames

        # Build column mapping
        col_map = {}
        for h in headers:
            normalized = normalize_column_name(h)
            if normalized:
                col_map[normalized] = h

        for row in reader:
            case = {}
            for standard, original in col_map.items():
                case[standard] = str(row.get(original, "")).strip()
            case["case_type"] = classify_case(case)
            cases.append(case)

    return cases


def parse_excel(filepath):
    """Parse Excel test case file."""
    if not HAS_OPENPYXL:
        print("Error: openpyxl is required for Excel files. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    cases = []
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Find header row (usually first non-empty row)
    headers = None
    header_row_idx = None
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=False), 1):
        cells = [str(c.value).strip() if c.value else "" for c in row]
        if any(cells):
            headers = cells
            header_row_idx = idx
            break

    if not headers:
        print("Error: No headers found in Excel file", file=sys.stderr)
        sys.exit(1)

    # Build column mapping
    col_map = {}
    for i, h in enumerate(headers):
        normalized = normalize_column_name(h)
        if normalized:
            col_map[normalized] = i

    # Parse data rows
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=False):
        case = {}
        for standard, col_idx in col_map.items():
            val = row[col_idx].value if col_idx < len(row) else None
            case[standard] = str(val).strip() if val else ""

        # Skip completely empty rows
        if not any(case.values()):
            continue

        case["case_type"] = classify_case(case)
        cases.append(case)

    wb.close()
    return cases


def main():
    parser = argparse.ArgumentParser(description="Parse test case files into structured JSON")
    parser.add_argument("input", help="Path to Excel (.xlsx) or CSV file")
    parser.add_argument("--output", "-o", help="Output JSON file path (default: stdout)")
    args = parser.parse_args()

    filepath = Path(args.input)
    if not filepath.exists():
        print(f"Error: File not found: {filepath}", file=sys.stderr)
        sys.exit(1)

    suffix = filepath.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        cases = parse_excel(filepath)
    elif suffix == ".csv":
        cases = parse_csv(filepath)
    else:
        print(f"Error: Unsupported file format: {suffix}", file=sys.stderr)
        sys.exit(1)

    result = {
        "source_file": str(filepath),
        "total_cases": len(cases),
        "cases_by_type": {
            "api": len([c for c in cases if c["case_type"] == "api"]),
            "webui": len([c for c in cases if c["case_type"] == "webui"]),
            "mixed": len([c for c in cases if c["case_type"] == "mixed"]),
        },
        "cases": cases,
    }

    if args.output:
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)
        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"Parsed {len(cases)} test cases -> {args.output}")
        print(f"  API: {result['cases_by_type']['api']}, WebUI: {result['cases_by_type']['webui']}, Mixed: {result['cases_by_type']['mixed']}")
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
